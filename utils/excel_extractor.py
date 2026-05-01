from io import BytesIO


def extract_text_from_excel(file_bytes: bytes) -> str:
    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    parts = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        all_rows = []
        header_row = None
        header_idx = None

        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            cells = [str(c).strip() if c is not None else '' for c in row]
            if not any(c for c in cells):
                continue

            # Detect the header row: contains price-related keywords
            if header_idx is None:
                joined = '\t'.join(cells).lower()
                if any(kw in joined for kw in ('price', 'rate', 'cost', 'total', 'service', 'unit')):
                    header_row = cells
                    header_idx = row_idx
                    all_rows.append(f"[HEADERS row {row_idx}]: " + '\t'.join(cells))
                    continue

            all_rows.append('\t'.join(cells))

        if all_rows:
            parts.append(f"=== Sheet: {sheet_name} ===\n" + '\n'.join(all_rows))

    wb.close()
    return '\n\n'.join(parts)

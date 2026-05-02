"""
Proposal blueprint - brief form, budget view, Excel export
"""
import uuid
import math
import json
from datetime import datetime
from io import BytesIO

from flask import Blueprint, render_template, request, jsonify, redirect, url_for, send_file
from models import db, Hotel, Service, ServiceCategory, Currency, Proposal
from utils.brief_extractor import extract_text_from_brief
from utils.ai_extractor import extract_from_brief_text, suggest_program, update_itinerary

proposal_bp = Blueprint('proposal', __name__, template_folder='templates')


# ── helpers ──────────────────────────────────────────────────────────────────

def _get_rate(currency_code):
    """Return target currency rate. Prices stored in USD (rate=1.0)."""
    target = Currency.query.filter_by(code=currency_code).first()
    return target.custom_rate if target else 1.0


def _usd_to(price_usd, target_rate):
    """Convert USD price to target currency."""
    return price_usd * target_rate


def _nights(arrival, departure):
    if arrival and departure and departure > arrival:
        return (departure - arrival).days
    return 1


def _build_budget(proposal):
    """Return structured budget dict for template and Excel. Prices stored in USD."""
    target_rate = _get_rate(proposal.currency_code)
    nights = _nights(proposal.arrival_date, proposal.departure_date)
    pax = proposal.pax

    hotel_ids = [int(x) for x in (proposal.hotels_selected or '').split(',') if x.strip()]
    service_ids = [int(x) for x in (proposal.services_selected or '').split(',') if x.strip()]

    hotels = Hotel.query.filter(Hotel.id.in_(hotel_ids)).all() if hotel_ids else []
    services = Service.query.filter(Service.id.in_(service_ids)).all() if service_ids else []

    hotel_rows = []
    for h in hotels:
        single_usd = h.rate_single_aed * nights   # field stores USD despite name
        twin_usd = h.rate_twin_aed * nights
        hotel_rows.append({
            'name': h.name,
            'location': h.location,
            'stars': h.stars,
            'website_url': h.website_url or '',
            'nights': nights,
            'rate_single_usd': h.rate_single_aed,
            'rate_twin_usd': h.rate_twin_aed,
            'total_single': _usd_to(single_usd, target_rate),
            'total_twin': _usd_to(twin_usd, target_rate),
        })

    service_rows = []
    for s in services:
        unit = s.unit or 'per person'
        if unit == 'per person':
            qty = pax
        elif unit == 'per day':
            qty = nights
        elif unit == 'per room':
            qty = math.ceil(pax / 2)
        else:
            qty = 1
        total_usd = s.price_aed * qty   # field stores USD despite name
        service_rows.append({
            'id': s.id,
            'name': s.name,
            'category': s.category.name if s.category else '',
            'price_usd': s.price_aed,
            'unit': unit,
            'qty': qty,
            'total': _usd_to(total_usd, target_rate),
        })

    services_total = sum(r['total'] for r in service_rows)
    hotels_total_single = sum(r['total_single'] for r in hotel_rows)
    hotels_total_twin = sum(r['total_twin'] for r in hotel_rows)

    return {
        'currency': proposal.currency_code,
        'nights': nights,
        'pax': pax,
        'hotels': hotel_rows,
        'services': service_rows,
        'services_total': services_total,
        'hotels_total_single': hotels_total_single,
        'hotels_total_twin': hotels_total_twin,
        'grand_total_single': hotels_total_single + services_total,
        'grand_total_twin': hotels_total_twin + services_total,
    }


# ── routes ────────────────────────────────────────────────────────────────────

@proposal_bp.route('/brief', methods=['GET'])
def brief_form():
    hotels = Hotel.query.filter_by(active=True).order_by(Hotel.name).all()
    services = Service.query.filter_by(active=True).order_by(Service.name).all()
    currencies = Currency.query.filter_by(active=True).all()
    return render_template('proposal/brief.html',
                           hotels=hotels, services=services, currencies=currencies)


@proposal_bp.route('/brief/parse', methods=['POST'])
def parse_brief():
    text = None
    if 'file' in request.files and request.files['file'].filename:
        f = request.files['file']
        text = extract_text_from_brief(f.read(), f.filename)
    elif request.form.get('text'):
        text = request.form.get('text')

    if not text or not text.strip():
        return jsonify({'error': 'No content provided'}), 400

    result = extract_from_brief_text(text)
    if 'error' in result:
        return jsonify({'error': result['error']}), 500
    return jsonify(result)


@proposal_bp.route('/brief/suggest', methods=['POST'])
def brief_suggest():
    data = request.get_json()
    if not data:
        return jsonify({'error': 'No data provided'}), 400

    nights = data.get('nights', 3)

    services = Service.query.filter_by(active=True).all()
    services_list = [
        {
            'id': s.id,
            'name': s.name,
            'category': s.category.name if s.category else 'Other',
            'price': s.price_aed,
            'unit': s.unit or 'per person',
            'description': s.description or '',
        }
        for s in services
        if s.category and s.category.name != 'Accommodation'
    ]

    if not services_list:
        return jsonify({'error': 'No services in database yet. Add services first.'}), 400

    result = suggest_program(
        group_type=data.get('group_type', ''),
        industry=data.get('industry', ''),
        pax=int(data.get('pax', 1)),
        nights=int(nights),
        special_requests=data.get('special_requests', ''),
        services=services_list,
    )

    if 'error' in result:
        return jsonify({'error': result['error']}), 500

    # Enrich itinerary items with service names for the editor
    svc_map = {s['id']: s for s in services_list}
    for day in result.get('itinerary', []):
        for item in day.get('items', []):
            sid = item.get('service_id')
            item['service_name'] = svc_map.get(sid, {}).get('name', 'Unknown')
            item['category'] = svc_map.get(sid, {}).get('category', '')

    # Collect all unique service_ids from itinerary
    result['service_ids'] = list({
        item['service_id']
        for day in result.get('itinerary', [])
        for item in day.get('items', [])
        if item.get('service_id')
    })

    return jsonify(result)


@proposal_bp.route('/brief/submit', methods=['POST'])
def brief_submit():
    def _parse_date(s):
        if not s:
            return None
        try:
            return datetime.strptime(s, '%Y-%m-%d')
        except ValueError:
            return None

    hotel_ids = ','.join(request.form.getlist('hotel_ids'))
    service_ids = ','.join(request.form.getlist('service_ids'))
    itinerary_raw = request.form.get('itinerary_json', '')

    proposal = Proposal(
        id=str(uuid.uuid4()),
        company_name=request.form.get('company_name', 'Unknown'),
        group_type=request.form.get('group_type', ''),
        industry=request.form.get('industry', ''),
        pax=int(request.form.get('pax', 1)),
        arrival_date=_parse_date(request.form.get('arrival_date')),
        departure_date=_parse_date(request.form.get('departure_date')),
        currency_code=request.form.get('currency', 'USD'),
        hotels_selected=hotel_ids,
        services_selected=service_ids,
        special_requests=request.form.get('special_requests', ''),
        itinerary_json=itinerary_raw if itinerary_raw else None,
        status='draft',
    )
    db.session.add(proposal)
    db.session.commit()
    return redirect(url_for('proposal.budget_view', proposal_id=proposal.id))


def _parse_itinerary(proposal):
    try:
        return json.loads(proposal.itinerary_json) if proposal.itinerary_json else []
    except Exception:
        return []


@proposal_bp.route('/proposal/<proposal_id>')
def budget_view(proposal_id):
    proposal = Proposal.query.get_or_404(proposal_id)
    budget = _build_budget(proposal)
    itinerary = _parse_itinerary(proposal)
    return render_template('proposal/budget.html', proposal=proposal, budget=budget, itinerary=itinerary)


@proposal_bp.route('/proposal/<proposal_id>/update-itinerary', methods=['POST'])
def update_itinerary_route(proposal_id):
    proposal = Proposal.query.get_or_404(proposal_id)
    data = request.get_json()
    change_request = data.get('change_request', '').strip()
    if not change_request:
        return jsonify({'error': 'No change request provided'}), 400

    current = _parse_itinerary(proposal)
    services = Service.query.filter_by(active=True).all()
    services_list = [
        {'id': s.id, 'name': s.name,
         'category': s.category.name if s.category else 'Other'}
        for s in services
    ]

    result = update_itinerary(current, change_request, services_list)
    if 'error' in result:
        return jsonify({'error': result['error']}), 500

    proposal.itinerary_json = json.dumps(result['itinerary'])
    db.session.commit()
    return jsonify({'itinerary': result['itinerary']})


@proposal_bp.route('/proposal/<proposal_id>/excel')
def export_excel(proposal_id):
    proposal = Proposal.query.get_or_404(proposal_id)
    budget = _build_budget(proposal)
    itinerary = _parse_itinerary(proposal)
    output = _generate_excel(proposal, budget, itinerary)
    filename = f"Proposal_{proposal.company_name.replace(' ', '_')}.xlsx"
    return send_file(output, as_attachment=True,
                     download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ── Excel generation ──────────────────────────────────────────────────────────

PINK = 'FF33CC'       # table header fill (matches original)
YELLOW = 'FFFF00'     # optional rows

def _generate_excel(proposal, budget, itinerary=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)          # remove default sheet; we add per hotel below
    wb.calculation.calcMode = 'auto'
    wb.calculation.fullCalcOnLoad = True

    cur = budget['currency']
    tgt_r = _get_rate(proposal.currency_code)
    pax = proposal.pax
    nights = budget['nights']

    def px(color):
        return PatternFill('solid', fgColor=color)

    def thin_border():
        s = Side(style='thin', color='CCCCCC')
        return Border(left=s, right=s, top=s, bottom=s)

    def _price(usd_val):
        return round(_usd_to(usd_val, tgt_r), 2)

    period_str = ''
    if proposal.arrival_date and proposal.departure_date:
        period_str = (f"{proposal.arrival_date.strftime('%d.%m.%Y')} to "
                      f"{proposal.departure_date.strftime('%d.%m.%Y')}")

    contact = 'Contact INSIDERS Tourism: Ahmed Shelleh, ahmed@Insiders-uae.com, Mob: +971 50 3434428 ; Tel : +971 445 39 822'
    currency_name = {'USD': 'US Dollars', 'EUR': 'Euro', 'AED': 'UAE Dirhams', 'GBP': 'British Pounds'}.get(cur, cur)

    notes = [
        '1. Rates include all local taxes and are subject to change until final confirmation.',
        '2. Nothing of the above is booked or blocked until deposit is received.',
        '3. Positions marked as "Option" are not included in the grand total.',
        '4. In case of changes in the number of participants, rates may be revised.',
    ]

    # ── One sheet per hotel ──
    hotels = budget['hotels']
    services = budget['services']

    if not hotels:
        # If no hotels selected, create one sheet with all services
        hotels = [{'name': 'Proposal', 'location': '', 'stars': 0,
                   'nights': nights, 'rate_single_aed': 0, 'rate_twin_aed': 0,
                   'total_single': 0, 'total_twin': 0}]

    for h in hotels:
        ws = wb.create_sheet(title=(h['name'][:28] + '  '))

        # Column widths (match original)
        for col, w in zip('ABCDEFGH', [24.44, 38.44, 7.44, 12.0, 28.66, 17.11, 52.44, 9.11]):
            ws.column_dimensions[col].width = w

        # ── Header block ──
        ws.cell(3, 4, 'Hotel Name')
        ws.cell(3, 5, h['name']).font = Font(bold=True, size=11)
        ws.cell(4, 4, 'Hotel website:')
        if h.get('website_url'):
            ws.cell(4, 5, h['website_url'])
        ws.cell(5, 1, f'Period: {period_str}').font = Font(bold=True)
        ws.cell(6, 1, f'Type of the group: {proposal.group_type or ""}')
        ws.cell(7, 1, f'Amount of pax: {pax}')
        ws.cell(8, 1, f'Destination: UAE - {h["location"] or "Dubai"}')
        ws.cell(9, 1, contact).font = Font(bold=True)
        ws.cell(10, 1, f'Currency - {currency_name}').font = Font(bold=True)

        # ── Table header (row 11) ──
        headers = ['Timing', 'Service', 'Unit, No', 'Frequency of units',
                   f'Price per unit, {cur}', f'Total, {cur}', 'Comments']
        aligns = ['center', 'left', 'center', 'center', 'center', 'center', 'center']
        for col, (hdr, aln) in enumerate(zip(headers, aligns), 1):
            c = ws.cell(11, col, hdr)
            c.font = Font(bold=True, size=11)
            c.fill = px(PINK)
            c.alignment = Alignment(horizontal=aln, vertical='center', wrap_text=True)
            c.border = thin_border()
        ws.row_dimensions[11].height = 29.4

        row = 12

        # ── Hotel rooms section ──
        ws.cell(row, 1, h['name']).font = Font(bold=True)
        row += 1

        single_price = _price(h['rate_single_usd'])
        twin_price = _price(h['rate_twin_usd'])

        # Twin row
        twin_row = row
        ws.cell(row, 2, 'Room Double/Twin').font = Font(bold=True)
        ws.cell(row, 2).alignment = Alignment(horizontal='left', wrap_text=True)
        ws.cell(row, 3, 0).alignment = Alignment(horizontal='center')
        ws.cell(row, 4, nights).alignment = Alignment(horizontal='center')
        ws.cell(row, 5, twin_price).alignment = Alignment(horizontal='center')
        ws.cell(row, 6, f'=C{row}*D{row}*E{row}').alignment = Alignment(horizontal='center')
        ws.row_dimensions[row].height = 28.8
        row += 1

        # Single row
        single_row = row
        ws.cell(row, 2, 'Room Single').font = Font(bold=True)
        ws.cell(row, 2).alignment = Alignment(horizontal='left', wrap_text=True)
        ws.cell(row, 3, pax).alignment = Alignment(horizontal='center')
        ws.cell(row, 4, nights).alignment = Alignment(horizontal='center')
        ws.cell(row, 5, single_price).alignment = Alignment(horizontal='center')
        ws.cell(row, 6, f'=C{row}*D{row}*E{row}').alignment = Alignment(horizontal='center')
        ws.row_dimensions[row].height = 28.8
        row += 1

        # Hotel subtotal row
        hotel_total_row = row
        if proposal.arrival_date:
            ws.cell(row, 1, proposal.arrival_date.strftime('%d.%m.%Y')).font = Font(bold=True)
        ws.cell(row, 6, f'=SUM(F{twin_row}:F{single_row})').font = Font(bold=True)
        row += 1

        # ── Services section ──
        # Build lookup dicts for quick access
        svc_by_id = {s['id']: s for s in services}
        svc_by_name = {s['name'].lower(): s for s in services}

        def _write_svc_row(ws, row, s, timing='', note=''):
            unit = s['unit']
            if unit == 'per person':
                qty, freq = pax, 1
            elif unit == 'per day':
                qty, freq = 1, nights
            elif unit == 'per room':
                qty, freq = math.ceil(pax / 2), 1
            else:
                qty, freq = 1, 1
            ws.cell(row, 1, timing).alignment = Alignment(horizontal='center')
            ws.cell(row, 2, s['name'])
            ws.cell(row, 2).alignment = Alignment(horizontal='left', wrap_text=True)
            ws.cell(row, 3, qty).alignment = Alignment(horizontal='center')
            ws.cell(row, 4, freq).alignment = Alignment(horizontal='center')
            ws.cell(row, 5, _price(s['price_usd'])).alignment = Alignment(horizontal='center')
            ws.cell(row, 6, f'=C{row}*D{row}*E{row}').alignment = Alignment(horizontal='center')
            ws.cell(row, 7, note or s['category'])
            ws.row_dimensions[row].height = 28.8
            return row

        svc_rows = []
        written_ids = set()

        if itinerary:
            for day in itinerary:
                # Day header row
                day_label = day.get('label') or f"Day {day['day']}"
                c = ws.cell(row, 2, f"Day {day['day']}  —  {day_label}")
                c.font = Font(bold=True, size=11)
                c.fill = px('D9E1F2')
                c.alignment = Alignment(horizontal='left')
                ws.merge_cells(f'A{row}:G{row}')
                ws.row_dimensions[row].height = 20
                row += 1

                for item in day.get('items', []):
                    sid = item.get('service_id')
                    sname = (item.get('service_name') or item.get('name') or '').lower()
                    s = svc_by_id.get(sid) or svc_by_name.get(sname)
                    if s:
                        _write_svc_row(ws, row, s,
                                       timing=item.get('timing', ''),
                                       note=item.get('note', ''))
                        svc_rows.append(row)
                        written_ids.add(s['id'])
                        row += 1
                    elif item.get('service_name') or item.get('name'):
                        # Service mentioned in itinerary but not in DB — write as text
                        name = item.get('service_name') or item.get('name') or ''
                        ws.cell(row, 1, item.get('timing', '')).alignment = Alignment(horizontal='center')
                        ws.cell(row, 2, name).alignment = Alignment(horizontal='left', wrap_text=True)
                        ws.cell(row, 7, item.get('note', ''))
                        ws.row_dimensions[row].height = 28.8
                        svc_rows.append(row)
                        row += 1

            # Write any selected services not mentioned in itinerary
            remaining = [s for s in services if s['id'] not in written_ids]
            if remaining:
                c = ws.cell(row, 2, 'Additional Services')
                c.font = Font(bold=True, size=11)
                c.fill = px('D9E1F2')
                ws.merge_cells(f'A{row}:G{row}')
                ws.row_dimensions[row].height = 20
                row += 1
                for s in remaining:
                    _write_svc_row(ws, row, s)
                    svc_rows.append(row)
                    row += 1
        else:
            # No itinerary — flat list as before
            for s in services:
                _write_svc_row(ws, row, s)
                svc_rows.append(row)
                row += 1

        # Services subtotal
        svc_total_row = row
        if svc_rows:
            ws.cell(row, 6, f'=SUM(F{svc_rows[0]}:F{svc_rows[-1]})').font = Font(bold=True)
        row += 1

        # ── Totals block ──
        total_row = row
        ws.cell(row, 2, f'Total Expenses {cur}').font = Font(bold=True)
        ws.cell(row, 2).fill = px(PINK)
        ws.cell(row, 2).alignment = Alignment(horizontal='left')
        total_f = f'=F{hotel_total_row}+F{svc_total_row}'
        ws.cell(row, 6, total_f).font = Font(bold=True)
        ws.cell(row, 6).fill = px(PINK)
        ws.cell(row, 6).alignment = Alignment(horizontal='center')
        row += 1

        ws.cell(row, 2, 'Grand Total').font = Font(bold=True)
        ws.cell(row, 2).fill = px(PINK)
        ws.cell(row, 6, f'=SUM(F{total_row}:F{total_row})').font = Font(bold=True)
        ws.cell(row, 6).fill = px(PINK)
        ws.cell(row, 6).alignment = Alignment(horizontal='center')
        row += 1

        ws.cell(row, 2, f'Total per 1 pax, {cur} based on {pax}-pax').font = Font(bold=True)
        ws.cell(row, 2).fill = px(PINK)
        ws.cell(row, 6, f'=F{row-1}/{pax}').font = Font(bold=True)
        ws.cell(row, 6).fill = px(PINK)
        ws.cell(row, 6).alignment = Alignment(horizontal='center')
        row += 3

        # ── Important notes ──
        ws.cell(row, 1, 'Important notes:').font = Font(bold=True)
        row += 1
        for note in notes:
            ws.cell(row, 1, note)
            row += 1

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


